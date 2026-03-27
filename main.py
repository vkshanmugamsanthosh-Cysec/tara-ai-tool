import sys
import json
import pandas as pd
from datetime import datetime
from risk_engine import compute_risk

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import landscape, A4

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Load Asset Library ─────────────────────────────
with open("asset_library.json") as f:
    asset_library = json.load(f)

# ── Inputs ─────────────────────────────────────────
ecu = sys.argv[1]
interface = sys.argv[2]

# ── Dynamic File Naming (CRITICAL FIX) ─────────────
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
excel_file = f"tara_{ecu}_{interface}_{timestamp}.xlsx"
pdf_file = f"tara_{ecu}_{interface}_{timestamp}.pdf"

# ── Generate TARA Data ─────────────────────────────
data = []
threat_id = 1

for asset in asset_library.get(ecu, []):
    for prop in asset["security_properties"]:
        for stride in [
            "Spoofing",
            "Tampering",
            "Denial of Service",
            "Information Disclosure",
            "Elevation of Privilege"
        ]:

            impact, feasibility, score, level = compute_risk(asset["type"], interface)

            data.append({
                "Threat ID": f"TH_{ecu}_{threat_id:03d}",
                "ECU": ecu,
                "Asset": asset["name"],
                "Asset Type": asset["type"],
                "Security Property": prop,
                "STRIDE": stride,
                "Impact": impact,
                "Feasibility": feasibility,
                "Risk Score": score,
                "Risk Level": level
            })

            threat_id += 1

df = pd.DataFrame(data)

# ── Excel Generation ───────────────────────────────
with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:

    # Sheet 1: TARA Results
    df.to_excel(writer, sheet_name="TARA Results", index=False)

    # Sheet 2: Standards / Scoring
    standards = pd.DataFrame({
        "Parameter": [
            "Safety Impact (S)",
            "Financial Impact (F)",
            "Operational Impact (O)",
            "Privacy Impact (P)",
            "Final Impact",
            "",
            "Attack Feasibility",
            "Expertise",
            "Knowledge",
            "Window of Opportunity",
            "Equipment",
            "",
            "Risk Score",
            "High",
            "Medium",
            "Low"
        ],
        "Description": [
            "Risk of injury or death",
            "Financial loss",
            "Operational disruption",
            "Data exposure",
            "MAX(S, F, O, P)",
            "",
            "Sum of feasibility factors",
            "Skill required",
            "System knowledge",
            "Time window",
            "Tools required",
            "",
            "Impact × Feasibility",
            ">= 20",
            "10–19",
            "< 10"
        ],
        "Range": [
            "0–3",
            "0–3",
            "0–3",
            "0–3",
            "0–3",
            "",
            "0–12",
            "1–3",
            "1–3",
            "1–3",
            "1–3",
            "",
            "0–36",
            "",
            "",
            ""
        ]
    })

    standards.to_excel(writer, sheet_name="Standards", index=False)

# ── Excel Styling ──────────────────────────────────
wb = load_workbook(excel_file)

# Style Sheet 1
ws1 = wb["TARA Results"]

header_fill = PatternFill(start_color="003366", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for col in ws1.columns:
    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
    ws1.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

ws1.freeze_panes = "A2"

for row in ws1.iter_rows(min_row=2):
    risk_cell = row[-1]
    if risk_cell.value == "High":
        risk_cell.fill = PatternFill(start_color="C00000", fill_type="solid")
        risk_cell.font = Font(color="FFFFFF", bold=True)
    elif risk_cell.value == "Medium":
        risk_cell.fill = PatternFill(start_color="FF8C00", fill_type="solid")
        risk_cell.font = Font(color="FFFFFF", bold=True)
    elif risk_cell.value == "Low":
        risk_cell.fill = PatternFill(start_color="375623", fill_type="solid")
        risk_cell.font = Font(color="FFFFFF", bold=True)

# Style Sheet 2
ws2 = wb["Standards"]

for cell in ws2[1]:
    cell.fill = PatternFill(start_color="1F3864", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)

for col in ws2.columns:
    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
    ws2.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

wb.save(excel_file)

# ── PDF Generation ─────────────────────────────────
styles = getSampleStyleSheet()

doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4))
content = []

content.append(Paragraph(f"TARA Report — {ecu} ECU ({interface})", styles["Title"]))
content.append(Spacer(1, 20))

table_data = [df.columns.tolist()] + df.values.tolist()

table = Table(table_data, repeatRows=1)
table.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE", (0, 0), (-1, -1), 7),
    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
]))

content.append(table)
doc.build(content)

# ── Output ─────────────────────────────────────────
print(f"Generated: {excel_file}")
print(f"Generated: {pdf_file}")